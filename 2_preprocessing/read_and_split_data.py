# -*- coding: utf-8 -*-
"""
Read and split data for image recognition project.

This script handles data preprocessing, including:
- Reading and resizing images
- Splitting data into train/val/test sets
- Balancing datasets
- Saving processed images to directories

"""

import cv2
import os
import shutil
import pandas as pd
import numpy as np
from pathlib import Path
from sklearn.model_selection import train_test_split
from imblearn.under_sampling import RandomUnderSampler
from imblearn.over_sampling import RandomOverSampler
from tqdm import tqdm
import multiprocessing
from functools import partial
import sys

import logging
logging.basicConfig(level=logging.INFO)

from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Add project root to path for config import
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config import (
    BASE_DIR, DATA_DIR, PREPROCESSING_DIR, MAPPING_FILE,
    get_imputed_patches_path, validate_paths, setup_environment
)

# Setup environment and validate paths
setup_environment(verbose=False)

# Configuration dictionary
CONFIG = {
    'image_size': (224, 224),
    'random_seed': 424242,  # 424242 used for all downstream runs, 1, 3 for sensitivity analysis
    'processing_methods': ['no_masking'],  # ['random_dataset', 'black', 'random_image', 'no_masking']
    'thresholds': [0.2],
    'partition_type': 'patient',  # Alternatives are 'WSI' or 'patch'
    'balancing_strategy': 'No balancing',  # Do during training
    'delete_existing_dirs': True,
    'window_size': (120, 120),  # Original patch size
}

def get_data_path(processing_method, threshold):
    """
    Generate the data path based on processing method and threshold.
    
    Args:
        processing_method: Imputation method ('black', 'random_image', 'random_dataset', 'no_masking')
        threshold: Tissue threshold (0.2, 0.3, ..., 0.8)
    
    Returns:
        Full path to the processed dataset directory
    """
    return str(get_imputed_patches_path(processing_method, threshold, CONFIG['window_size']))

def read_img(img_path, size):
    """
    Read and resize a single image patch.
    
    Args:
        img_path: Path to image file
        size: Target size tuple (width, height)
    
    Returns:
        Resized image array (BGR format from OpenCV)
    """
    img = cv2.imread(str(img_path))
    return cv2.resize(img, size)

def list_images(main_path, size):
    """
    Load image data and labels from directory structure.
    
    Scans directory structure to find all image patches and extracts labels
    from directory names (Inside/Outside). Merges with patient metadata
    from mapping.csv file.
    
    Args:
        main_path: Base path containing Inside/ and Outside/ directories
        size: Image size (used for validation, not resizing here)
    
    Returns:
        DataFrame with columns: Full_path, Image, Label, tiff, and patient metadata
    """
    folders = [a for a in os.listdir(main_path) if os.path.isdir(os.path.join(main_path, a)) and 'dir' not in a]
    dirs = [os.path.join(main_path, f) for f in folders]

    images = []
    for i, dir_path in enumerate(dirs):
        for count, img_path in enumerate(Path(dir_path).glob('**/*.png'), 1):
            images.append([img_path, os.path.basename(img_path), folders[i].split('_')[0]])
            if count % 1000 == 0:
                print(f'Loading data from dir {folders[i]}: {count}')

    images_df = pd.DataFrame(images, columns=['Full_path', 'Image', 'Label'])
    images_df['tiff'] = [name[:name.find('_')] for name in images_df['Image']]

    # Join subject info - use config path
    if not MAPPING_FILE.exists():
        raise FileNotFoundError(
            f"Mapping file not found at: {MAPPING_FILE}\n"
            f"Please ensure 1_data/mapping.csv exists."
        )
    mapping = pd.read_csv(str(MAPPING_FILE), sep=';')
    return images_df.merge(mapping, on='tiff', how='left')

def create_individual_partitions(images, train_size=0.6, val_size=0.2, test_size=0.2, output_path=None, threshold=None, overwrite=False, random_seed=42):
    """
    Create patient-level partitions for train/validation/test sets.
    
    This function implements patient-level data splitting to prevent data leakage.
    All patches from the same patient are assigned to the same split. The function
    maintains stratification by COPD group to ensure balanced distribution.
    
    Partitions are saved to Excel file for reproducibility. If partitions already
    exist for threshold 0.2, they are reused for other thresholds.
    
    Args:
        images: DataFrame with image paths and patient metadata
        train_size: Proportion of data for training (default: 0.6)
        val_size: Proportion of data for validation (default: 0.2)
        test_size: Proportion of data for testing (default: 0.2)
        output_path: Directory to save partition Excel file
        threshold: Tissue threshold (used for file naming)
        overwrite: Whether to overwrite existing partitions
        random_seed: Random seed for reproducibility (default: 42, use 424242 for manuscript)
    
    Returns:
        DataFrame with columns: patient, partition (train/val/test)
    
    Note:
        For manuscript reproduction, random_seed must be 424242.
        Split ratios: 60% train, 20% validation, 20% test.
    """
    # Use config path if output_path not specified
    if output_path is None:
        output_path = str(PREPROCESSING_DIR)
    
    excel_path = os.path.join(output_path, 'patient_partitions.xlsx')

    if not overwrite:
        # Read existing partitions from the Excel file
        if os.path.exists(excel_path):
            partitions_df = pd.read_excel(excel_path, sheet_name='Partitions')
            base_partitions = partitions_df[partitions_df['threshold'] == 0.2][['patient', 'partition']]
            if not base_partitions.empty:
                print(f"Using existing partitions from threshold 0.2 for threshold {threshold}")
                
                # Calculate patch counts and WSI counts for the current threshold
                patch_counts = images.groupby('patient')['Image'].count()
                wsi_counts = images.groupby('patient')['tiff'].nunique()
                
                # Create a copy of the Excel file for the new threshold
                new_excel_path = os.path.join(output_path, f'patient_partitions_{threshold}.xlsx')
                shutil.copy(excel_path, new_excel_path)
                
                # Update the new Excel file with threshold-specific information
                with pd.ExcelWriter(new_excel_path, engine='openpyxl', mode='a') as writer:
                    # Update Partitions sheet
                    partitions = base_partitions.copy()
                    partitions['threshold'] = threshold
                    partitions[f'patches_{threshold}'] = partitions['patient'].map(patch_counts)
                    partitions[f'wsi_count_{threshold}'] = partitions['patient'].map(wsi_counts)
                    partitions.to_excel(writer, sheet_name=f'Partitions_{threshold}', index=False)
                    
                    # Update Patient Details sheet
                    if 'Patient Details' in pd.ExcelFile(new_excel_path).sheet_names:
                        patient_details = pd.read_excel(new_excel_path, sheet_name='Patient Details')
                        patient_details[f'patches_{threshold}'] = patient_details['patient_id'].map(patch_counts)
                        patient_details[f'wsi_count_{threshold}'] = patient_details['patient_id'].map(wsi_counts)
                        patient_details.to_excel(writer, sheet_name=f'Patient Details_{threshold}', index=False)
                
                print(f"Created new Excel file with information for threshold {threshold}: {new_excel_path}")
                
                return partitions
            else:
                print(f"No existing partitions found for threshold 0.2. Creating new partitions.")
        else:
            print("Excel file not found. Creating new partitions.")

    # If we reach here, we need to create new partitions
    # We base the patient partitioning on the smallest threshold
    elif threshold == 0.2 and overwrite:
        # Create DataFrame with patient info and their patch counts
        patient_info = (images.groupby('patient')
                       .agg({
                           'Image': 'count',
                           'copd_group': 'first'
                       })
                       .rename(columns={'Image': 'patch_count'}))
        
        total_patches = len(images)
        train_target = int(total_patches * train_size)
        val_target = int(total_patches * val_size)
        test_target = total_patches - train_target - val_target

        # Initialize containers for each split
        train_patients, val_patients, test_patients = [], [], []
        train_count, val_count, test_count = 0, 0, 0

        # Stratify by copd_group
        for copd_group in patient_info['copd_group'].unique():
            copd_patients = patient_info[patient_info['copd_group'] == copd_group]
            copd_total = copd_patients['patch_count'].sum()
            
            # Calculate target counts for this copd_group
            copd_test_target = int(copd_total * test_size)
            copd_val_target = int(copd_total * val_size)
            
            # Shuffle patients within this copd_group
            copd_patients = copd_patients.sample(frac=1, random_state=CONFIG['random_seed'])
            
            current_count = 0
            for patient, row in copd_patients.iterrows():
                if current_count < copd_test_target:
                    test_patients.append(patient)
                    test_count += row['patch_count']
                elif current_count < (copd_test_target + copd_val_target):
                    val_patients.append(patient)
                    val_count += row['patch_count']
                else:
                    train_patients.append(patient)
                    train_count += row['patch_count']
                current_count += row['patch_count']

        partitions = pd.DataFrame({
            'patient': train_patients + val_patients + test_patients,
            'partition': ['train'] * len(train_patients) + ['val'] * len(val_patients) + ['test'] * len(test_patients)
        })

        # Print summary statistics
        print("\nOverall split statistics:")
        print(f"Train: {train_count} patches ({train_count/total_patches:.2%})")
        print(f"Validation: {val_count} patches ({val_count/total_patches:.2%})")
        print(f"Test: {test_count} patches ({test_count/total_patches:.2%})")
        
        print("\nCopd distribution in splits:")
        for split, patients in [('Train', train_patients), ('Val', val_patients), ('Test', test_patients)]:
            copd_dist = patient_info.loc[patients, 'copd_group'].value_counts()
            print(f"\n{split} split copds:")
            print(copd_dist)


    if output_path and overwrite:
        excel_path = os.path.join(output_path, 'patient_partitions.xlsx')
        
        if overwrite or not os.path.exists(excel_path):
            wb = Workbook()
        else:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path)

        # First tab: Partitions
        if 'Partitions' not in wb.sheetnames:
            ws1 = wb.create_sheet('Partitions')
        else:
            ws1 = wb['Partitions']
        
        # Add threshold column to partitions
        partitions['threshold'] = threshold
        
        # Append new data to the existing sheet
        if ws1.max_row == 1:  # If sheet is empty, add headers
            ws1.append(['patient', 'partition', 'threshold'])
        for r in dataframe_to_rows(partitions, index=False, header=False):
            ws1.append(r)

        # Second tab: Detailed Patient Info
        if 'Patient Details' not in wb.sheetnames:
            ws2 = wb.create_sheet('Patient Details')
        else:
            ws2 = wb['Patient Details']

        # Create the detailed patient info dataframe
        unique_patients = np.unique(images['patient'])
        patient_details = pd.DataFrame(index=unique_patients)
        patient_details['train'] = patient_details.index.isin(train_patients).astype(int)
        patient_details['val'] = patient_details.index.isin(val_patients).astype(int)
        patient_details['test'] = patient_details.index.isin(test_patients).astype(int)
        
        # Count image patches for each patient in each partition
        for partition in ['train', 'val', 'test']:
            partition_patients = locals()[f'{partition}_patients']
            patient_details[f'{partition}_patches_{threshold}'] = images[images['patient'].isin(partition_patients)]['patient'].value_counts()
        
        # Fill NaN values with 0 for patch counts
        patient_details = patient_details.fillna(0)
        
        # Convert patch counts to integers
        for col in [f'{partition}_patches_{threshold}' for partition in ['train', 'val', 'test']]:
            patient_details[col] = patient_details[col].astype(int)
        
        # Add patient details to the second worksheet
        if ws2.max_row == 1:  # If sheet is empty, add headers
            ws2.append(['patient_id', 'train', 'val', 'test'] + 
                       [f'{partition}_patches_{threshold}' for partition in ['train', 'val', 'test']] +
                       [f'{partition}_wsi_count' for partition in ['train', 'val', 'test']])
        
        # Calculate WSI counts for each patient in each partition
        for partition in ['train', 'val', 'test']:
            partition_patients = locals()[f'{partition}_patients']
            patient_details[f'{partition}_wsi_count'] = images[images['patient'].isin(partition_patients)].groupby('patient')['tiff'].nunique()
        
        # Fill NaN values with 0 for WSI counts
        patient_details = patient_details.fillna(0)
        
        # Convert WSI counts to integers
        for col in [f'{partition}_wsi_count' for partition in ['train', 'val', 'test']]:
            patient_details[col] = patient_details[col].astype(int)
        
        for patient, row in patient_details.iterrows():
            existing_row = next((r for r in ws2.iter_rows(min_row=2, values_only=True) if r[0] == patient), None)
            if existing_row:
                row_index = existing_row[0]
                for col, value in enumerate(row.tolist(), start=2):
                    ws2.cell(row=row_index, column=col, value=value)
            else:
                ws2.append([patient] + row.tolist())
        
        # Save the Excel file
        wb.save(excel_path)
        print(f"Patient partitions and details saved to {excel_path}")
    
    return partitions


def split_data(images, partition_type, processing_method, threshold, partitions=None, test_size=0.2, val_size=0.2):
    """
    Split image data into train/validation/test sets.
    
    Supports three splitting strategies:
    - 'patient': Patient-level splitting (prevents data leakage, used in manuscript)
    - 'WSI': WSI-level splitting (all patches from same WSI in same split)
    - 'patch': Patch-level splitting (random, may cause data leakage)
    
    Args:
        images: DataFrame with image paths and metadata
        partition_type: 'patient', 'WSI', or 'patch'
        processing_method: Imputation method (for logging)
        threshold: Tissue threshold (for logging)
        partitions: DataFrame with patient partitions (required for 'patient' type)
        test_size: Proportion for test set
        val_size: Proportion for validation set
    
    Returns:
        Tuple of (train_indices, val_indices, test_indices, summary_table)
        summary_table: DataFrame with patch counts per patient per split
    """
    print("Starting data split...")
    
    mask_org = ~images['Image'].str.contains('_rot.png')
    images_org = images[mask_org]
    indices_org = images_org.index.tolist()

    if partition_type == "patch":
        print("Performing non-independent split...")
        with tqdm(total=len(indices_org), desc="Train-test split") as pbar:
            ind_train, ind_test = train_test_split(indices_org,
                                                   test_size=test_size,
                                                   random_state=42,
                                                   stratify=images_org['copd_group'],
                                                   shuffle=True)
            pbar.update(len(indices_org))
        print(f"Train-test split complete. Train size: {len(ind_train)}, Test size: {len(ind_test)}")
        
        with tqdm(total=len(ind_train), desc="Train-validation split") as pbar:
            ind_train, ind_val = train_test_split(ind_train,
                                                  test_size=val_size/(1-test_size),
                                                  random_state=42,
                                                  stratify=images.loc[ind_train, 'copd_group'],
                                                  shuffle=True)
            pbar.update(len(ind_train))
        print(f"Train-validation split complete. Train size: {len(ind_train)}, Validation size: {len(ind_val)}")

    elif partition_type == "WSI":
        print("Performing WSI-level split...")
        # Get unique WSIs and their labels
        wsi_data = images_org.groupby('tiff')['copd_group'].first().reset_index()
        
        if wsi_data.empty:
            raise ValueError("No WSI data found in the dataset")

        # Print first rows of wsi_data to inspect the data
        print("\nFirst rows of WSI data:")
        print(wsi_data.head())
        print(f"\nTotal number of WSIs: {len(wsi_data)}")
        print(f"COPD group distribution:\n{wsi_data['copd_group'].value_counts()}\n")
        
        # Split WSIs into train/test
        wsi_train, wsi_test = train_test_split(
            wsi_data['tiff'],
            test_size=test_size,
            random_state=42,
            stratify=wsi_data['copd_group'],  # Changed from Label to copd_group
            shuffle=True
        )
        
        # Split train WSIs into train/val
        wsi_train, wsi_val = train_test_split(
            wsi_train,
            test_size=val_size/(1-test_size),
            random_state=42,
            stratify=wsi_data[wsi_data['tiff'].isin(wsi_train)]['copd_group'],  # Changed from Label to copd_group
            shuffle=True
        )
        # Get indices for each split based on WSI assignments
        ind_train = images_org[images_org['tiff'].isin(wsi_train)].index.tolist()
        ind_val = images_org[images_org['tiff'].isin(wsi_val)].index.tolist()
        ind_test = images_org[images_org['tiff'].isin(wsi_test)].index.tolist()
        
        print(f"Split complete using WSIs:")
        print(f"Train size: {len(ind_train)} patches from {len(wsi_train)} WSIs")
        print(f"Validation size: {len(ind_val)} patches from {len(wsi_val)} WSIs")
        print(f"Test size: {len(ind_test)} patches from {len(wsi_test)} WSIs")

    elif partition_type == "patient":
        print("Performing independent split using partitions...")
        # Create masks for each partition using the partitions dataframe
        train_mask = images_org['patient'].isin(partitions[partitions['partition'] == 'train']['patient'])
        val_mask = images_org['patient'].isin(partitions[partitions['partition'] == 'val']['patient'])
        test_mask = images_org['patient'].isin(partitions[partitions['partition'] == 'test']['patient'])

        # Get indices for each partition
        ind_train = images_org[train_mask].index.tolist()
        ind_val = images_org[val_mask].index.tolist()
        ind_test = images_org[test_mask].index.tolist()

        print(f"Split complete using partitions:")
        print(f"Train size: {len(ind_train)} ({len(ind_train)/len(images_org):.2%})")
        print(f"Validation size: {len(ind_val)} ({len(ind_val)/len(images_org):.2%})")
        print(f"Test size: {len(ind_test)} ({len(ind_test)/len(images_org):.2%})")

    print(f"Split complete. Train size: {len(ind_train)}, Validation size: {len(ind_val)}, Test size: {len(ind_test)}")
    
    # Modified code for creating summary table
    all_patients = sorted(list(set(images['patient'].unique())))  # Get all unique patients
    train_counts = images.loc[ind_train]['patient'].value_counts()
    val_counts = images.loc[ind_val]['patient'].value_counts()
    test_counts = images.loc[ind_test]['patient'].value_counts()

    summary_table = pd.DataFrame(index=all_patients)
    summary_table['Train'] = summary_table.index.map(train_counts).fillna(0).astype(int)
    summary_table['Validation'] = summary_table.index.map(val_counts).fillna(0).astype(int)
    summary_table['Test'] = summary_table.index.map(test_counts).fillna(0).astype(int)
    summary_table.reset_index(inplace=True)
    summary_table.rename(columns={'index': 'Patient'}, inplace=True)

    print("\nSummary of patches per individual:")
    print(summary_table)

    return ind_train, ind_val, ind_test, summary_table

def balance_dataset(images, ind_train, balancing_strategy):
    print(f"Balancing dataset using strategy: {balancing_strategy}")
    
    # Create mappings for quick lookups
    filename_to_index = {filename: index for index, filename in enumerate(images.Image)}
    index_to_filename = {index: filename for filename, index in filename_to_index.items()}
    index_to_label = {index: label for index, label in zip(images.index, images.Label)}
    
    # Logging to check what is happening
    logging.info(f"Max index in index_to_filename: {max(index_to_filename.keys())}")
    logging.info(f"Max index in ind_train before filtering: {max(ind_train)}")
    
    # Filter out indices that are not in index_to_filename
    ind_train = [i for i in ind_train if i in index_to_filename]
    
    logging.info(f"Max index in ind_train after filtering: {max(ind_train)}")
    

    if balancing_strategy == 'No balancing':
        print("No balancing applied.")
        fn_train = [index_to_filename[i] for i in tqdm(ind_train, desc="Collecting train images")]
    elif balancing_strategy == 'Ext_oversample':
        print("Applying external oversampling...")
        rot180_train_imgs = [index_to_filename[i].replace('.png','_rot.png') 
                             for i in tqdm(ind_train, desc="Finding rotated images") 
                             if index_to_label[i] == 'Inside']
        fn_train = [index_to_filename[i] for i in tqdm(ind_train, desc="Collecting train images")] + rot180_train_imgs
        print(f"Added {len(rot180_train_imgs)} rotated images.")
    elif balancing_strategy == 'RandomSampler':
        print("Applying random sampling...")
        sample_strat = RandomOverSampler(sampling_strategy='minority') if balancing_strategy == 'Oversample' else RandomUnderSampler(sampling_strategy='majority')
        ind_new, _ = sample_strat.fit_resample(np.arange(len(images)).reshape(-1, 1), images.Label)
        ind_new = ind_train[ind_new].squeeze()
        fn_train = [index_to_filename[i] for i in tqdm(ind_new, desc="Collecting sampled images")]
        print(f"Sampled dataset size: {len(fn_train)}")
    else:
        print(f"Unknown balancing strategy: '{balancing_strategy}'. No balancing applied.")
        fn_train = [index_to_filename[i] for i in tqdm(ind_train, desc="Collecting train images")]

    # Use the mapping to get indices quickly
    result = [filename_to_index[fn] for fn in tqdm(fn_train, desc="Creating final index list")]
    print(f"Balancing complete. Final dataset size: {len(result)}")
    return result

def resize_data(main_path, size):
    """Resize all images in the given directory."""
    for folder in os.listdir(main_path):
        if os.path.isdir(os.path.join(main_path, folder)):
            dir_path = os.path.join(main_path, folder)
            for count, img_path in enumerate(Path(dir_path).glob('**/*.png'), 1):
                img = read_img(img_path, size)
                cv2.imwrite(str(img_path), img)
                if count % 1000 == 0:
                    print(f'Resizing data from dir {folder}: {count}')

def create_directories(output_data_path, delete_existing=False):
    """Create necessary directories for train, validation, and test data."""
    dirs = ['train_dir', 'val_dir', 'test_dir']
    classes = ['Outside', 'Inside']
    
    for dir_name in dirs:
        dir_path = os.path.join(output_data_path, dir_name)
        if delete_existing and os.path.exists(dir_path):
            try:
                shutil.rmtree(dir_path)
            except FileNotFoundError:
                # If the directory doesn't exist, just continue
                pass
        os.makedirs(dir_path, exist_ok=True)
        for class_name in classes:
            os.makedirs(os.path.join(dir_path, class_name), exist_ok=True)

def save_and_resize_images(images, indices, target_dir, output_data_path, size):
    """Save and resize images to the specified directory."""
    for count, i in enumerate(indices, 1):
        img_path = images.Full_path[i]
        img_name = images.Image[i]
        label = images.Label[i]
        target_path = os.path.join(target_dir, label, img_name)

        # Read, resize, and save the image
        img = read_img(img_path, size)
        cv2.imwrite(target_path, img)

        if count % 1000 == 0:
            print(f'Saving and resizing images: {count}')

def process_single_configuration(processing_method, threshold):
    print(f"Processing: {processing_method}, Threshold: {threshold}")
    
    # Use configuration from CONFIG dictionary
    data_path = get_data_path(processing_method, threshold)
    
    # Create directories
    create_directories(data_path, delete_existing=CONFIG['delete_existing_dirs'])

    # Process data
    print('Listing all images')
    images = list_images(data_path, CONFIG['image_size'])

    print('Creating individual partitions')
    if CONFIG['partition_type'] == "patient":
        partitions = create_individual_partitions(
            images, 
            output_path=str(PREPROCESSING_DIR), 
            threshold=threshold, 
            overwrite=False, 
            random_seed=CONFIG['random_seed']
        )
    else:
        partitions = None
    
    print('Splitting dataset')
    ind_train, ind_val, ind_test, summary_table = split_data(images, CONFIG['partition_type'], processing_method, threshold, partitions)
    print('Balancing dataset')
    ind_train_new = balance_dataset(images, ind_train, CONFIG['balancing_strategy'])

    # Save images
    datasets = [
        ('train', ind_train_new),
        ('val', ind_val),
        ('test', ind_test)
    ]

    for dataset_name, indices in datasets:
        target_dir = os.path.join(data_path, f'{dataset_name}_dir')
        print(f'Saving and resizing {dataset_name} images')
        save_and_resize_images(images, indices, target_dir, data_path, CONFIG['image_size'])

    # Write summary table as CSV
    summary_csv_path = os.path.join(data_path, 'summary_table.csv')
    summary_table.to_csv(summary_csv_path, index=False)
    print(f"Summary table saved to {summary_csv_path}")

    print(f"Completed processing for {processing_method}, Threshold: {threshold}")

def main():
    configurations = [(method, threshold) for method in CONFIG['processing_methods'] for threshold in CONFIG['thresholds']]
    
    if len(configurations) > 1:
        print("Running parallel processing for multiple configurations")
        num_cores = multiprocessing.cpu_count()
        print(f"Using {num_cores} CPU cores for parallel processing")
        with multiprocessing.Pool(num_cores) as pool:
            pool.starmap(process_single_configuration, configurations)
    else:
        print("Running single configuration")
        process_single_configuration(*configurations[0])

if __name__ == "__main__":
    main()
